import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from azure.devops.connection import Connection
from azure.devops.v7_1.work_item_tracking.models import Wiql
from msrest.authentication import BasicAuthentication
from datetime import datetime
import traceback
import re
from bs4 import BeautifulSoup
from dotenv import load_dotenv
import os

# Load environment variables
load_dotenv()

# Configuration
ADO_ORG_URL = "https://dev.azure.com/skype"  # Replace with your organization
PROJECT = "SPOOL"  # Replace with your project name
# Create a new workbook
workbook = openpyxl.Workbook()
# List of AreaPaths
area_paths = ['SPOOL\\PSTN', 'SPOOL\\Call Automation\\Signaling', 'SPOOL\\Call Automation\\Media', 'SPOOL\\Graph\\Calling', 'SPOOL\\Job Router', 'SPOOL\\Graph\\Application Management', 'SPOOL\\Calling\\Captions', 'SPOOL\\Calling\\Recording']
ADO_PAT =  os.getenv('PAT') #"4bZyoofO8FM7oJ0Q6JAXZQMvEaAwZSQxKDt2Zy2wkGnUgFWMuXQKJQQJ99ALACAAAAAAArohAAASAZDO4VSr"  # Replace with your ADO personal access token

# Define color fills
grey_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Light grey
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
orange_fill = PatternFill(start_color="FFD9AD", end_color="FFD9AD", fill_type="solid")  # Orange
red_fill = PatternFill(start_color="FF7070", end_color="FF7070", fill_type="solid")    # Light red
no_fill = PatternFill(fill_type=None)

def get_ado_connection():
    credentials = BasicAuthentication('', ADO_PAT)
    connection = Connection(base_url=ADO_ORG_URL, creds=credentials)
    return connection

# Query to get all features with Funding Status 'Funded'
def get_features_with_funding_status(area_path):
    connection = get_ado_connection()
    wit_client = connection.clients_v7_1.get_work_item_tracking_client()
    
    wiql = Wiql(
        query=f"""
        SELECT [System.Id], [System.Tags]
        FROM WorkItems
        WHERE [System.WorkItemType] = 'Feature'
        AND [System.AreaPath] == '{area_path}'
        AND NOT [System.State] IN ('Removed', 'Closed', 'Completed')
        AND [Skype.Funding] IN ('Partially Funded', 'Funded')
        """
    )
    
    result = wit_client.query_by_wiql(wiql).work_items

    # List of tags to exclude
    exclude_tags = ['sfi', 'security wave', 'securitywave', 'ic3 horizontal', 'ic3horizontal']

    filtered_features = []
    for item_ref in result:
        item = wit_client.get_work_item(item_ref.id)
        tags = item.fields.get('System.Tags', '').lower()
        if not any(exclude_tag in tags for exclude_tag in exclude_tags):
            filtered_features.append(item)

    return [item.id for item in filtered_features]

# Function to get detailed information of a work item by ID
def get_work_item_details(work_item_id):
    connection = get_ado_connection()
    wit_client = connection.clients_v7_1.get_work_item_tracking_client()
    return wit_client.get_work_item(work_item_id, expand='All')

# Function to retrieve user stories linked to a feature
def get_tasks_for_feature(feature_id):
    connection = get_ado_connection()
    wit_client = connection.clients_v7_1.get_work_item_tracking_client()

    wiql = Wiql(
        query=f"""
        SELECT [System.Id]
        FROM WorkItems
        WHERE [System.WorkItemType] = 'User Story'
        AND [System.Tags] CONTAINS 'PV-Status-Tracker'
        AND [System.Parent] = {feature_id}
        """
    )
    
    result = wit_client.query_by_wiql(wiql).work_items

    if result:
        story_id = result[0].id
        wiql = Wiql(
            query=f"""
            SELECT [System.Id]
            FROM WorkItems
            WHERE [System.WorkItemType] = 'Task'
            AND [System.Parent] = {story_id}
            """
        )

        result = wit_client.query_by_wiql(wiql).work_items
        return [wit_client.get_work_item(item.id) for item in result]
    
    return []

# Function to extract the end date of the iteration path
def get_iteration_end_date(work_item, state):
    iteration_path = work_item.fields.get('System.IterationPath', None)
    if iteration_path:
        end_date = extract_date_from_iteration_path(iteration_path, state)
        return end_date if end_date else "N/A"
    return "N/A"

def extract_date_from_iteration_path(iteration_path_str, state):
    try:
        # Extract the date substring between 'Ends' and ')'
        date_part = iteration_path_str.split("Ends")[-1].split(")")[0].strip()
        current_year = datetime.now().year

        # Convert date_part (e.g., 'Nov 24') to a date
        date_obj = datetime.strptime(date_part, '%b %d')
        date_obj = date_obj.replace(year=current_year)

        # Check if the date has already passed this year
        if state != 'Resolved' and state != 'Closed' and state != 'Removed' and date_obj < datetime.now():
            date_obj = date_obj.replace(year=current_year + 1)

        return date_obj.strftime('%b %d, %Y')
    except (IndexError, ValueError):
        return None

def get_date_state_changed(work_item, target_state="In Progress"):
    connection = get_ado_connection()
    wit_client = connection.clients_v7_1.get_work_item_tracking_client()
    
    # Get all revisions of the work item
    revisions = wit_client.get_revisions(work_item.id, expand='Fields')

    # Iterate over revisions to find when the state changed to the target state
    for revision in revisions:
        state = revision.fields.get('System.State')
        if state == target_state:
            return revision.fields.get('System.ChangedDate')

    return "N/A"  # Return None if the state change wasn't found

def reformat_date(date_str):
    if date_str == "N/A":
        return date_str
    
    # Parse the date string to a datetime object
    date_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")

    # Format the datetime object to the desired format
    return date_obj.strftime("%b %d, %Y")

def check_pattern(string):
    # Define the regex pattern
    pattern = r"SPOOL\\CY\d{4}-H[1-2]\\CY\d{4}-Q[1-4]"
    
    # Check if the string matches the pattern
    if re.fullmatch(pattern, string):
        return True
    else:
        return False

def days_until_iteration_path_end(input_string):
    try:
        # Extract the date string from the input
        start_index = input_string.find("(Ends ") + 6
        end_index = input_string.find(")", start_index)
        date_str = input_string[start_index:end_index]

        # Get the current date
        today = datetime.now()

        # Parse the date string to a datetime object
        # Use current year if the extracted date's month is greater than or equal to current month, otherwise use next year
        year = today.year if datetime.strptime(date_str, '%b %d').month >= today.month else today.year + 1
        date_obj = datetime.strptime(date_str + ' ' + str(year), '%b %d %Y')

        # Calculate the number of days from today to the date object
        delta = date_obj - today

        # Return the number of days
        return max(1, delta.days)  # Ensure we don't return negative values

    except (ValueError, IndexError):
        # If the date string is not properly formatted or can't be extracted, return 0
        return -1
    
def track_feature_completion_ratio(work_item):
    connection = get_ado_connection()
    wit_client = connection.clients_v7_1.get_work_item_tracking_client()

    # Get the iteration path of the feature
    feature_iteration_path = work_item.fields['System.IterationPath']
    days_until_feature_completion = days_until_iteration_path_end(feature_iteration_path)

    # No Iteration Path Set
    if (days_until_feature_completion == -1 and not check_pattern(feature_iteration_path)):
         return grey_fill
    
    child_wiql_query_string = """
        SELECT
            [System.Id],
            [System.WorkItemType],
            [System.Title],
            [System.State],
            [Microsoft.VSTS.Scheduling.StoryPoints]
        FROM WorkItemLinks
        WHERE
            ([Source].[System.WorkItemType] = 'Feature' AND [Source].[System.Id] = {work_item_id})
            AND ([Target].[System.WorkItemType] = 'User Story')
            AND ([Target].[System.State] <> 'Removed')
            AND ([Target].[System.State] <> 'Closed')
            AND ([System.Links.LinkType] = 'System.LinkTypes.Hierarchy-Forward')
        MODE (Recursive)
    """.format(work_item_id=work_item.id)
    
    # Create the Wiql object using the formatted query string
    child_wiql_query = Wiql(query=child_wiql_query_string)

    # Get child items of the feature
    child_wiql_results = wit_client.query_by_wiql(child_wiql_query).work_item_relations

        # Extract the work item IDs from the query results
    child_item_ids = [relation.target.id for relation in child_wiql_results if relation.target]

    # Get the work items
    child_items = wit_client.get_work_items(ids=child_item_ids)
    
    # Sum the story points of all user stories
    total_story_points = 0
    for child_work_item_reference in child_items:
        child_work_item_id = child_work_item_reference.id
        child_work_item = wit_client.get_work_item(child_work_item_id, fields=['Microsoft.VSTS.Scheduling.StoryPoints'])
        story_points = child_work_item.fields.get('Microsoft.VSTS.Scheduling.StoryPoints', 0)
        total_story_points += story_points

    status_perc = (total_story_points / days_until_feature_completion) * 100

    if status_perc > 99:
        return red_fill
    elif status_perc > 89 and status_perc <= 99:
        return orange_fill
    elif status_perc > 49 and status_perc <= 89:
        return green_fill
    else:
        return no_fill
    
def fill_and_wrap_entire_row(sheet, row_num, fill, max_col=None):
    """Apply a fill color to an entire row.
    
    Args:
        sheet: The worksheet object.
        row_num: The row number to fill.
        fill: A PatternFill object with the desired color.
        max_col: The maximum column to fill (defaults to sheet.max_column).
    """
    if max_col is None:
        max_col = sheet.max_column  # Automatically determine the range
    for col in range(1, max_col + 1):  # Iterate through all columns in the row
        sheet.cell(row=row_num, column=col).fill = fill
        sheet.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

# Function to create Excel workbook and populate data
def create_excel_report(features, area_path):
    # Create a new sheet for the current AreaPath, replace \ with _ & remove the SPOOL\ prefix from the iteration path in the sheet title
    sheet = workbook.create_sheet(title=area_path.replace("\\", "_")[6:])

    # Set column titles
    columns = [
        'Feature', 'Requirements lock date', 'Design completion date', 'UX Design ETA',
        'Privacy review', 'Security reviews', 'Code start date', 'Azure Review',
        'Code completion date', 'Testing', 'Bug fixing', 'Deployment', 'Status Note', 'Blockers', 'Dependencies',
        'Action items with owners'
    ]
    for col_num, title in enumerate(columns, start=1):
        col_letter = get_column_letter(col_num)  # Get the column letter (e.g., 'A', 'B', etc.)
        sheet[f'{col_letter}1'] = title  # Set the title in the cell
        # Set the desired width for the column
        if col_letter == 'A':
            sheet.column_dimensions[col_letter].width = 30  
        elif col_letter == 'M':
            sheet.column_dimensions[col_letter].width = 40  
        else:
            sheet.column_dimensions[col_letter].width = 14  

    # Populate data rows
    for row_num, feature_id in enumerate(features, start=2):
        feature_details = get_work_item_details(feature_id)
        feature_name = feature_details.fields.get('System.Title', 'N/A')
        feature_url = f"{ADO_ORG_URL}/{PROJECT}/_workitems/edit/{feature_id}"
        functional_spec_eta = reformat_date(feature_details.fields.get('Custom.FunctionalSpecETA', 'N/A'))
        code_start_eta = reformat_date(feature_details.fields.get('Microsoft.VSTS.Scheduling.StartDate', 'N/A'))
        deployment_eta =get_iteration_end_date(feature_details, feature_details.fields.get('System.State')) 
        # Extract the status note
        status_note = feature_details.fields.get("Skype.StatusNotes", "")
        # Regex pattern to match the first <div>...</div>
        match = re.search(r"<div>.*?</div>", status_note)
        if match:
            status_note = match.group(0)
        
        # Add data to the row
        sheet[f'A{row_num}'].hyperlink = feature_url
        sheet[f'A{row_num}'].value = feature_name
        sheet[f'B{row_num}'].value = functional_spec_eta
        sheet[f'G{row_num}'].value = code_start_eta
        sheet[f'L{row_num}'].value = deployment_eta
        # Get everything up to the first newline
        plain_text = BeautifulSoup(status_note, "html.parser").get_text()
        sheet[f'M{row_num}'].value = plain_text

        tasks = get_tasks_for_feature(feature_id)

        # Iterate through user stories to populate specific fields
        for task in tasks:
            tags = task.fields.get('System.Tags', '').split(';')
            iteration_end_date = get_iteration_end_date(task, task.fields.get('System.State'))

            if 'PV-Tracker-Design' in tags:
                sheet[f'C{row_num}'].value = iteration_end_date
            if 'PV-Tracker-UX' in tags:
                sheet[f'D{row_num}'].value = iteration_end_date
            if 'PV-Tracker-Privacy' in tags:
                sheet[f'E{row_num}'].value = iteration_end_date
            if 'PV-Tracker-Security' in tags:
                sheet[f'F{row_num}'].value = iteration_end_date
            if 'PV-Tracker-Azure-Review' in tags:
                sheet[f'H{row_num}'].value = iteration_end_date
            if 'PV-Tracker-Testing' in tags:
                sheet[f'J{row_num}'].value = iteration_end_date
            if 'PV-Tracker-Bug-Fix' in tags:
                sheet[f'K{row_num}'].value = iteration_end_date
            if 'PV-Tracker-Codecomplete' in tags:
                sheet[f'I{row_num}'].value = iteration_end_date
        
        # Fill placeholder columns
        for col, value in zip(['N', 'O', 'P'], [''] * 4):
            sheet[f'{col}{row_num}'].value = value

        fill = track_feature_completion_ratio(feature_details)
        fill_and_wrap_entire_row(sheet=sheet, row_num=row_num, fill=fill)

    
    # workbook.save('projects_status_report.xlsx')

if __name__ == "__main__":
    try:
        for area_path in area_paths:
            feature_ids = get_features_with_funding_status(area_path)
            if feature_ids:
                create_excel_report(feature_ids, area_path)
                print(f"{area_path} sheet generated successfully!")
            else:
                print("No features found with Funding Status 'Funded'.")

        # Remove the default sheet created by openpyxl
        default_sheet = workbook['Sheet']
        workbook.remove(default_sheet)

        # Save the workbook
        workbook.save('projects_status_report.xlsx')
    except Exception as e:
        print(f"An error occurred: {e}")
        traceback.print_exc()